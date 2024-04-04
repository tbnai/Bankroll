import openpyxl
import os

# Load the WorkBook
wb = openpyxl.load_workbook('Gamble2024.xlsx')
sheet = wb.active

# Find the last non-empty row
last_row = sheet.max_row
while sheet.cell(row=last_row, column=1).value is None:
    last_row -= 1

# Input data: Date, Buy-In, Win/Lose, Amount
input_date = input("Enter the Date: ")
buy_in = float(input("Enter the Buy-In amount: "))
win_or_lose = input("Did you Win or Lose?:(W/L): ").upper()
amount = float(input("Enter the Amount: "))

# Get the last row's 'Total Win/Lose' and calculate new total
prev_total_win_lose = sheet.cell(row=last_row, column=5).value
if prev_total_win_lose is None:
    prev_total_win_lose = 0  # Initialize to 0 if no previous value

if win_or_lose == 'W':
    new_total_win_lose = prev_total_win_lose + amount
else:
    new_total_win_lose = prev_total_win_lose - amount

# Write the input data into the new row
new_row = last_row + 1
sheet.cell(row=new_row, column=1).value = input_date
sheet.cell(row=new_row, column=2).value = buy_in
sheet.cell(row=new_row, column=3).value = amount if win_or_lose == 'W' else None
sheet.cell(row=new_row, column=4).value = amount if win_or_lose == 'L' else None
sheet.cell(row=new_row, column=5).value = new_total_win_lose

# Save the updated workbook
wb.save('Gamble2024.xlsx')

# Opens the sheet after saving
os.system('start Gamble2024.xlsx')